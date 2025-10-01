
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
import os
from datetime import date
import requests
from bs4 import BeautifulSoup
import pandas as pd
import shutil

today = date.today()

def strip_words(s):
    if s == 'null':
        return s
    else:
        s = s.replace(',','.')
        result = re.sub("[^0-9.-]", "", s)
        if result != '':
            return re.sub("[^0-9.-]", "", s)
        else:
            return 'null'

def correct_price(filename):
    c = get_column('Prezzo',filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[c]) + 1):
        price_val = sheet[f"{c}{i}"].value
        if price_val == 'null':
            sheet[f"{c}{i}"] = 'null'
        elif price_val != "Prezzo su richiesta":
            if 'da' not in price_val:
                if len(price_val.split()) > 1:
                    price_val = price_val.split()[1]
            price_val = price_val.replace(",00", "")
            price_val = strip_words(price_val)
            sheet[f"{c}{i}"] = price_val
    workbook.save(filename=filename)

def correct_field(filename,name,custom_interaction):
    col = get_column(name, filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value
        if room_value != "null":
            if custom_interaction:
                if name == 'piano': # Piano
                    if room_value in ['Piano T', 'Piano R', 'Piano S', 'Piano S - T']:
                        room_value = room_value[-1]
                    else:
                        room_value = strip_words(room_value)
                elif name == 'spese condominio': # Spese condominiali
                    if room_value == 'Nessuna spesa condominiale':
                        room_value = '0'
                    else:
                        room_value = strip_words(room_value)
                elif name == 'Box, posti auto': # Posti auto
                    nums = room_value.split(',')
                    acc = 0
                    for n in nums:
                        acc += int(strip_words(n)) if strip_words(n) != '' else 0
                    room_value = str(acc)
            else:
                room_value = strip_words(room_value)
            sheet[f"{col}{i}"] = room_value
    workbook.save(filename=filename)

def correct_categorical_field(filename, name, yes_label, no_label):
    col = get_column(name, filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value

        if room_value == yes_label:
                room_value = 'Si'

        elif room_value == no_label:
                room_value = 'No'

        sheet[f"{col}{i}"] = room_value
    workbook.save(filename=filename)


def place_space(filename,name,to_replace):
    col = get_column(name, filename)
    custom_interaction = True
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value
        if room_value is not None:
            if custom_interaction:
                 room_value = room_value.replace('- Scarica capitolato', '')
            room_value = room_value.replace(to_replace,', ')
            if room_value == '':
                room_value = 'null'
            sheet[f"{col}{i}"] = room_value
    workbook.save(filename=filename)

def split_values(filename,name):
    col = get_column(name, filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value

        if name == 'Tipo':  #   tipologia
            if room_value == 'null' or room_value is None:
                sheet[f"{get_prev(col,3)}{i}"] = 'null'
                sheet[f"{get_prev(col,2)}{i}"] = 'null'
                sheet[f"{get_prev(col)}{i}"] = 'null'
            else:
                words = room_value.split(' | ')
                sheet[f"{get_prev(col,3)}{i}"] = words[0]
                if len(words) > 1 and 'Classe' in words[-1]:
                    sheet[f"{get_prev(col,2)}{i}"] = words[-1]
                else:
                    sheet[f"{get_prev(col,2)}{i}"] = 'null'
                if len(words) > 1:
                    result = ""
                    for word in words[1:len(words)]:
                        result += word 
                        if word != words[-1]:
                            result += ", "
                    sheet[f"{get_prev(col)}{i}"] = result
                else:
                    sheet[f"{get_prev(col)}{i}"] = 'null'
        elif name == 'Prestazione energetica del fabbricato': # Prestazione energetica del fabbricato
            if 'estivo' in room_value:
                sheet[f"{get_prev(col,2)}{i}"] = 'si'
            else:
                sheet[f"{get_prev(col,2)}{i}"] = 'no'
            
            if 'invernale' in room_value:
                sheet[f"{get_prev(col)}{i}"] = 'si'
            else:
                sheet[f"{get_prev(col)}{i}"] = 'no'
        elif name == 'Riscaldamento': #riscaldamento
            if room_value == 'null':
                sheet[f"{get_prev(col,2)}{i}"] = 'null'
                sheet[f"{get_prev(col)}{i}"] = 'null'
            else:
                words = room_value.split(', ')
                if len(words) == 1:
                    sheet[f"{get_prev(col,2)}{i}"] = words[0]
                    sheet[f"{get_prev(col)}{i}"] = 'null'
                else:
                    sheet[f"{get_prev(col,2)}{i}"] = words[0]
                    sheet[f"{get_prev(col)}{i}"] = room_value.replace(words[0] + ', ', '')
        elif name == 'Efficienza energetica': #efficienza energetica
            if room_value == 'null':
                sheet[f"{get_prev(col,2)}{i}"] = 'null'
                sheet[f"{get_prev(col)}{i}"] = 'null'
            else:
                if len(room_value) <= 2:
                    sheet[f"{get_prev(col,2)}{i}"] = room_value
                    sheet[f"{get_prev(col)}{i}"] = 'null'
                elif room_value[0] == 'A' and room_value[1].isnumeric():
                    sheet[f"{get_prev(col,2)}{i}"] = room_value[0:2]
                    sheet[f"{get_prev(col)}{i}"] = strip_words(room_value[2:])
                else:
                    sheet[f"{get_prev(col,2)}{i}"] = room_value[0]
                    sheet[f"{get_prev(col)}{i}"] = strip_words(room_value[1:])
                      
    workbook.save(filename=filename)

def add_nulls(filename,name):
    col = get_column(name,filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value
        words = room_value.split(', ') 
        if len(words) != 4:
            to_add = ''
            for _ in range(4 - len(words)):
                to_add += ', null'
                sheet[f"{col}{i}"] = room_value + to_add
    workbook.save(filename=filename)
        
def fix_class_imm(filename):
    col = get_column('Altri dettagli su tipologia',filename)
    col_class = get_column('Classe Immobile',filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value
        if 'Immobile di lusso' in room_value:
            sheet[f"{col_class}{i}"] = 'Immobile di lusso'
        words = room_value.split(', ')
        room_value = room_value.replace(words[-1],'')
        sheet[f"{col}{i}"] = room_value if room_value != '' else 'null'
    workbook.save(filename=filename)

def get_civic(filename):
    col = 'C'
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    for i in range(2, len(sheet[col]) + 1):
        room_value = sheet[f"{col}{i}"].value
        numero_civico = re.search(r'\b\d{1,4}[A-Za-z]?\b', room_value)
        if numero_civico is not None:
            val = numero_civico.group()
            if val != "0" and val != "00":
                sheet[f"D{i}"] = val
            else:
                sheet[f"D{i}"] = 'null'
        else:
            sheet[f"D{i}"] = 'null'
    workbook.save(filename=filename)

def get_column(column_name, file_path):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Iterate through the first row to find the column name
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == column_name:
            # Return the column letter
            return get_column_letter(col[0].column)
    
    # If the column name is not found, return None or raise an exception
    return None

def get_prev(column_letter, dec=1):
    # Convert the column letter to an index
    column_index = column_index_from_string(column_letter)
    
    # Check if there is a previous column
    if column_index > dec:
        previous_column_index = column_index - dec
        # Convert the previous column index back to a letter
        return get_column_letter(previous_column_index)
    else:
        # If there is no previous column (i.e., column A), return None or handle as needed
        return None

def insert_columns(file_path, start_column_letter, num_columns, column_names):
    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Convert the starting column letter to an index
    start_column_index = column_index_from_string(start_column_letter)
    
    # Insert the specified number of columns at the starting index
    sheet.insert_cols(start_column_index, num_columns)
    
    # Set the names of the new columns in the first row and make them bold
    bold_font = Font(bold=True)
    for i in range(num_columns):
        if i < len(column_names):
            cell = sheet.cell(row=1, column=start_column_index + i)
            cell.value = column_names[i]
            cell.font = bold_font
    
    # Save the workbook
    workbook.save(file_path)



def correct_excel(path):

    # Siccome la vecchia colonna 'Tipologia' deve essere eliminata, la rinominiamo per evitare che si vada ad attingere dalla nuova colonna inserita vuota.
    tip_col = get_column('Tipologia',path)
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    sheet[tip_col + '1'] = 'Tipo'
    workbook.save(filename=path)

    insert_columns(path, tip_col, 3, ['Tipologia', 'Classe Immobile', 'Altri dettagli su tipologia'])
    if get_column('Riscaldamento',path) is not None:
        insert_columns(path, get_column('Riscaldamento',path), 2, ['Tipologia riscaldamento', 'Alimentazione riscaldamento'])
        split_values(path,'Riscaldamento')
    # insert_columns(path, get_column('Efficienza energetica',path),2, ['Classe energetica', 'Efficienza energetica (numero)'])
    # insert_columns(path, get_column('Prestazione energetica del fabbricato',path), 2, ['Prestazione energetica estiva?', 'Prestazione energetica invernale?'])
    get_civic(path)                                                 # estrai il numero civico dagli indirizzi
    correct_price(path)                                             # rappresenta il prezzo come numero
    correct_categorical_field(path,'ascensore?','Ascensore','No Ascensore')  # Rappresenta ascensore con si o no
    correct_categorical_field(path,'balcone?','Balcone','null')            # Rappresenta balcone con si o no
    correct_categorical_field(path,'terrazzo?','Terrazzo','null')           # Rappresenta terrazzo con si o no
    place_space(path,'Citta-zona-via-civico',';')                                       # Separa le indicazioni sull'indirizzo con ', ' rispetto a ';'
    place_space(path,'altre caratteristiche',';')                                       # Separa le indicazioni sulle altre caratteristiche
    place_space(path,'Contratto','|')                                       # Separa le indicazioni sulle indicazioni del contratto
    correct_field(path,'n locali',False)                                   # Rappresenta il numero di locali come numero
    correct_field(path,'area',False)                                   # Rappresenta l'area come numero
    correct_field(path,'n bagni',False)                                   # Rappresenta il numero di bagni come numero
    correct_field(path,'piano',True)                                    # Rappresenta il piano come numero
    if get_column('spese condominio',path) is not None:
        correct_field(path,'spese condominio',True)                        # Rappresenta le spese condominiali come numero
    split_values(path,'Tipo')                                         # Separa le colonne della tipologia
    fix_class_imm(path)                                             # Sposta i valori nella posizione sbagliata
    if get_column('Box, posti auto', path) is not None:
        correct_field(path,'Box, posti auto',True)                                   # Rappresenta i posti auto come numero
    # split_values(path,'Prestazione energetica del fabbricato')                                        # Separa le colonne della prestazione energetica
    # Separa le colonne del riscaldamento
    correct_field(path,'Efficienza energetica (numero)',False)                                        # Separa le colonne dell'efficenza energetica
    add_nulls(path,'Citta-zona-via-civico')                                             # Aggiunge dei vuoti alla colonna della via

def get_list_of_features(path,feature):
    col = get_column(feature,path)
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    result = set()
    column_values = [c.value for c in sheet[col]]
    unnecessary = [feature, '', 'null']
    for val in column_values:
        if val != None:
            words = val.split(',')
            words = [w.lstrip(' ').rstrip(' ') for w in words]
        else:
            words = []
        result.update(words)
    for u in unnecessary:
        if u in result:
            result.remove(u)
    return result

def get_list_of_feature_samples(sample_folder,feature):
    result = set()
    for path in os.listdir(sample_folder):
        result = result | get_list_of_features(sample_folder + '/' + path,feature)
    return result

def write_list(path,feature,column,letter):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    cell = sheet[f"{letter}1"]
    cell.value = feature
    cell.font = Font(bold=True)
    for idx, value in enumerate(column, start=2):
        sheet[f"{letter}{idx}"] = value
    workbook.save(filename=path)

def generate_table_poss(path):
    tipologia = get_list_of_feature_samples('regioni','Tipologia')
    write_list(path,'Tipologia',tipologia,'A')

    contratto = get_list_of_feature_samples('regioni', 'contratto')
    write_list(path,'contratto',contratto,'B')

    altri_dett = get_list_of_feature_samples('regioni','Altri dettagli su tipologia')
    write_list(path,'Altri dettagli su tipologia',altri_dett,'C')

    riscaldamento = get_list_of_feature_samples('regioni','riscaldamento')
    write_list(path,'riscaldamento',riscaldamento,'D')
    
    altri_dett = get_list_of_feature_samples('regioni','altre caratteristiche')
    infissi = [s for s in altri_dett if 'Infissi' in s]
    esposizione = [s for s in altri_dett if 'Esposizione' in s]
    arredamento = [s for s in altri_dett if 'Arred' in s]
    write_list(path,'Infissi',infissi,'E')
    write_list(path,'Esposizione',esposizione,'F')
    altro = [x for x in altri_dett if x not in infissi and x not in esposizione and x not in arredamento]
    write_list(path,'Arredamento',arredamento,'G')
    write_list(path,'altre caratteristiche',altro,'H')

def get_cap(city):
    # Load the CSV file
    cap_df = pd.read_csv('gi_comuni_cap.csv', sep=';', dtype={'cap': str})
    
    # Filter the DataFrame for the specified city
    cap_list = cap_df.loc[cap_df['denominazione_ita'] == city, 'cap']
    
    # Convert the Series to a string of comma-separated values
    cap_list_str = ', '.join(cap_list.astype(str))
    
    return cap_list_str

def split_into_folders(path):
    if not os.path.exists(f"{path}/real estate"):
        os.makedirs(f"{path}/real estate")

    if not os.path.exists(f"{path}/def"):
        os.makedirs(f"{path}/def")
    
    if not os.path.exists(f"{path}/corretto"):
        os.makedirs(f"{path}/corretto")

    for p in os.listdir(path):
        if os.path.splitext(p)[1] == '.xlsx':
            if 'real estate' in p:
                shutil.move(f"{os.path.abspath(path)}/{p}",f"{os.path.abspath(path)}/real estate/{p}")
            elif '-Def.xlsx' in p:
                shutil.move(f"{os.path.abspath(path)}/{p}",f"{os.path.abspath(path)}/def/{p}")
            else:
                shutil.move(f"{os.path.abspath(path)}/{p}",f"{os.path.abspath(path)}/corretto/{p}")



def generate_definitive_table(input_path, output_path):

    workbook_in = load_workbook(filename=input_path)
    workbook_out = Workbook()

    sheet_in = workbook_in.active
    sheet_out = workbook_out.active

    gray = PatternFill(patternType='solid',fgColor='DDD9C4')
    cyan = PatternFill(patternType='solid',fgColor='009999')
    blue = PatternFill(patternType='solid',fgColor='4BACC6')
    green = PatternFill(patternType='solid',fgColor='92D050')
    orange = PatternFill(patternType='solid',fgColor='F79646')

    link_col = get_column('Link',input_path)
    area_info_col = get_column('Citta-zona-via-civico',input_path)
    address_col = get_column('Indirizzo',input_path)
    civ_col = get_column('Civico',input_path)
    area_col = get_column('area',input_path)
    type_col = get_column('Tipologia',input_path)
    year_col = get_column('Anno di costruzione',input_path)
    class_col = get_column('Classe Immobile',input_path)
    price_col = get_column('Prezzo',input_path)
    state_col = get_column('Stato',input_path)
    floor_col = get_column('piano',input_path)
    lift_col = get_column('ascensore?',input_path)
    features_col = get_column('altre caratteristiche',input_path)
    auto_col = get_column('Box, posti auto',input_path)
    balcony_col = get_column('balcone?',input_path)
    terrace_col = get_column('terrazzo?',input_path)
    disp_col = get_column('Disponibilità',input_path)
    spese_col = get_column('Spese condominio',input_path)
    contr_col = get_column('Contratto',input_path)
    proprieta_col = get_column('Altri dettagli su tipologia',input_path)
    warm_type_col = get_column('Tipologia riscaldamento',input_path)
    warm_alim_col = get_column('Alimentazione riscaldamento',input_path)
    class_ene_col = get_column('Classe energetica',input_path)
    ene_eff_col = get_column('Efficienza energetica (numero)', input_path)
    ene_cert_col = get_column('certificazione energetica', input_path)
    summer_col = get_column('Prestazione energetica estiva?', input_path)
    winter_col = get_column('Prestazione energetica invernale?', input_path)
    index_ene_col = get_column('Indice prest. energetica rinnovabile', input_path)
    coord_col = get_column('Coord',input_path)
    cantina_col = get_column('Cantina',input_path)
    garden_col = get_column('Giardino',input_path)
    disable_col = get_column('Accesso disabili',input_path)
    portineria_col = get_column('Servizio portineria',input_path)


    link_cell = sheet_out['A1']
    link_cell.value = 'Link'
    link_cell.font = Font(bold=True)
    link_cell.fill = gray

    date_cell = sheet_out['B1']
    date_cell.value = 'Data rilevamento'
    date_cell.font = Font(bold=True)
    date_cell.fill = gray

    city_cell = sheet_out['C1']
    city_cell.value = 'Città'
    city_cell.font = Font(bold=True)
    city_cell.fill = gray

    coord_cell = sheet_out['D1']
    coord_cell.value = 'Coordinate'
    coord_cell.font = Font(bold=True)
    coord_cell.fill = gray

    via_cell = sheet_out['E1']
    via_cell.value = 'Via'
    via_cell.font = Font(bold=True)
    via_cell.fill = gray

    civ_cell = sheet_out['F1']
    civ_cell.value = 'Civico'
    civ_cell.font = Font(bold=True)
    civ_cell.fill = gray

    area_cell = sheet_out['G1']
    area_cell.value = 'Superficie'
    area_cell.font = Font(bold=True)
    area_cell.fill = gray

    type_cell = sheet_out['H1']
    type_cell.value = 'Tipologia'
    type_cell.font = Font(bold=True)
    type_cell.fill = gray

    year_cell = sheet_out['I1']
    year_cell.value = 'Anno di costruzione'
    year_cell.font = Font(bold=True)
    year_cell.fill = gray

    class_cell = sheet_out['J1']
    class_cell.value = 'Classe Immobile'
    class_cell.font = Font(bold=True)
    class_cell.fill = gray

    price_cell = sheet_out['K1']
    price_cell.value = 'Prezzo'
    price_cell.font = Font(bold=True)
    price_cell.fill = gray

    state_cell = sheet_out['L1']
    state_cell.value = 'Stato di conservazione'
    state_cell.font = Font(bold=True)
    state_cell.fill = gray

    floor_cell = sheet_out['M1']
    floor_cell.value = 'Piano'
    floor_cell.font = Font(bold=True)
    floor_cell.fill = gray

    lift_cell = sheet_out['N1']
    lift_cell.value = 'Ascensore'
    lift_cell.font = Font(bold=True)
    lift_cell.fill = cyan

    access_cell = sheet_out['O1']
    access_cell.value = 'Accesso per disabili'
    access_cell.font = Font(bold=True)
    access_cell.fill = cyan

    auto_cell = sheet_out['P1']
    auto_cell.value = 'Posti Auto'
    auto_cell.font = Font(bold=True)
    auto_cell.fill = cyan

    bike_cell = sheet_out['Q1']
    bike_cell.value = 'Parcheggio Bici'
    bike_cell.font = Font(bold=True)
    bike_cell.fill = cyan

    balcony_cell = sheet_out['R1']
    balcony_cell.value = 'Balcone'
    balcony_cell.font = Font(bold=True)
    balcony_cell.fill = cyan

    terrace_cell = sheet_out['S1']
    terrace_cell.value = 'Terrazza'
    terrace_cell.font = Font(bold=True)
    terrace_cell.fill = cyan

    private_garden_cell = sheet_out['T1']
    private_garden_cell.value = 'Giardino Privato'
    private_garden_cell.font = Font(bold=True)
    private_garden_cell.fill = cyan

    public_garden_cell = sheet_out['U1']
    public_garden_cell.value = 'Giardino Comune'
    public_garden_cell.font = Font(bold=True)
    public_garden_cell.fill = cyan

    pool_cell = sheet_out['V1']
    pool_cell.value = 'Piscina'
    pool_cell.font = Font(bold=True)
    pool_cell.fill = cyan

    cant_cell = sheet_out['W1']
    cant_cell.value = 'Cantina'
    cant_cell.font = Font(bold=True)
    cant_cell.fill = cyan

    tav_cell = sheet_out['X1']
    tav_cell.value = 'Taverna'
    tav_cell.font = Font(bold=True)
    tav_cell.fill = cyan

    portiere_intera_cell = sheet_out['Y1']
    portiere_intera_cell.value = 'Portiere intera giornata'
    portiere_intera_cell.font = Font(bold=True)
    portiere_intera_cell.fill = cyan

    portiere_mezza_cell = sheet_out['Z1']
    portiere_mezza_cell.value = 'Portiere mezza giornata'
    portiere_mezza_cell.font = Font(bold=True)
    portiere_mezza_cell.fill = cyan

    reception_cell = sheet_out['AA1']
    reception_cell.value = 'Reception'
    reception_cell.font = Font(bold=True)
    reception_cell.fill = cyan

    disp_cell = sheet_out['AB1']
    disp_cell.value = 'Disponibilità'
    disp_cell.font = Font(bold=True)
    disp_cell.fill = blue

    spese_cell = sheet_out['AC1']
    spese_cell.value = 'Spese condominio'
    spese_cell.font = Font(bold=True)
    spese_cell.fill = blue

    contr_cell = sheet_out['AD1']
    contr_cell.value = 'Contratto'
    contr_cell.font = Font(bold=True)
    contr_cell.fill = blue

    proprieta_cell = sheet_out['AE1']
    proprieta_cell.value = 'Proprietà'
    proprieta_cell.font = Font(bold=True)
    proprieta_cell.fill = blue

    warm_type_cell = sheet_out['AF1']
    warm_type_cell.value = 'Tipologia riscaldamento'
    warm_type_cell.font = Font(bold=True)
    warm_type_cell.fill = green

    warm_alim_cell = sheet_out['AG1']
    warm_alim_cell.value = 'Alimentazione riscaldamento'
    warm_alim_cell.font = Font(bold=True)
    warm_alim_cell.fill = green

    warm_src_cell = sheet_out['AH1']
    warm_src_cell.value = 'Fonte riscaldamento'
    warm_src_cell.font = Font(bold=True)
    warm_src_cell.fill = green

    class_ene_cell = sheet_out['AI1']
    class_ene_cell.value = 'Classe energetica'
    class_ene_cell.font = Font(bold=True)
    class_ene_cell.fill = green

    ene_eff_cell = sheet_out['AJ1']
    ene_eff_cell.value = 'Efficienza energetica'
    ene_eff_cell.font = Font(bold=True)
    ene_eff_cell.fill = green

    summer_cell = sheet_out['AK1']
    summer_cell.value = 'Prestazione energetica estiva?'
    summer_cell.font = Font(bold=True)
    summer_cell.fill = green

    winter_cell = sheet_out['AL1']
    winter_cell.value = 'Prestazione energetica invernale?'
    winter_cell.font = Font(bold=True)
    winter_cell.fill = green

    index_ene_cell = sheet_out['AM1']
    index_ene_cell.value = 'Indice prest. energetica rinnovabile'
    index_ene_cell.font = Font(bold=True)
    index_ene_cell.fill = green
    
    inf_type_cell = sheet_out['AN1']
    inf_type_cell.value = 'Tipologia di infissi'
    inf_type_cell.font = Font(bold=True)
    inf_type_cell.fill = green

    inf_mat_cell = sheet_out['AO1']
    inf_mat_cell.value = 'Materiale infissi'
    inf_mat_cell.font = Font(bold=True)
    inf_mat_cell.fill = green

    fireplace_cell = sheet_out['AP1']
    fireplace_cell.value = 'Caminetto'
    fireplace_cell.font = Font(bold=True)
    fireplace_cell.fill = green

    exp_cell = sheet_out['AQ1']
    exp_cell.value = 'Esposizione'
    exp_cell.font = Font(bold=True)
    exp_cell.fill = green

    techno_cell = sheet_out['AR1']
    techno_cell.value = 'Fibra Ottica'
    techno_cell.font = Font(bold=True)
    techno_cell.fill = orange






    for i in range(2, sheet_in.max_row + 1):
        sheet_out['A'+ str(i)] = sheet_in[link_col + str(i)].value
        sheet_out['B'+ str(i)] = str(date.today())
        sheet_out['C'+ str(i)] = sheet_in[area_info_col + str(i)].value.split(',')[0]
        sheet_out['D'+ str(i)] = sheet_in[coord_col + str(i)].value
        sheet_out['E'+ str(i)] = sheet_in[area_info_col + str(i)].value.split(',')[2]
        sheet_out['F'+ str(i)] = sheet_in[civ_col + str(i)].value
        sheet_out['G'+ str(i)] = sheet_in[area_col + str(i)].value
        sheet_out['H'+ str(i)] = sheet_in[type_col + str(i)].value
        if year_col is not None:
            sheet_out['I'+ str(i)] = sheet_in[year_col + str(i)].value
        sheet_out['J'+ str(i)] = sheet_in[class_col + str(i)].value
        sheet_out['K'+ str(i)] = sheet_in[price_col + str(i)].value.replace('.','')
        sheet_out['L'+ str(i)] = sheet_in[state_col + str(i)].value
        sheet_out['M'+ str(i)] = sheet_in[floor_col + str(i)].value
        sheet_out['N'+ str(i)] = 'Si' if sheet_in[lift_col + str(i)].value == 'Si' else 'No'
        sheet_out['O' + str(i)] = 'Si' if sheet_in[disable_col + str(i)].value in ('Sì','Si') else 'No'
        if auto_col is not None:
            sheet_out['P' + str(i)] = sheet_in[auto_col + str(i)].value if sheet_in[auto_col + str(i)].value != 'null' else str(0)
        sheet_out['Q' + str(i)] = 'Si' if 'Parcheggio bici' in sheet_in[features_col + str(i)].value else 'No'
        sheet_out['R' + str(i)] = 'Si' if sheet_in[balcony_col + str(i)].value == 'Si' else 'No'
        sheet_out['S' + str(i)] = 'Si' if sheet_in[terrace_col + str(i)].value == 'Si' else 'No'
        sheet_out['T' + str(i)] = 'Si' if 'Giardino privato' == sheet_in[garden_col + str(i)].value or 'Giardino privato e comune' == sheet_in[garden_col + str(i)].value else 'No'
        sheet_out['U' + str(i)] = 'Si' if 'Giardino comune' == sheet_in[garden_col + str(i)].value or 'Giardino privato e comune' == sheet_in[garden_col + str(i)].value else 'No'
        sheet_out['V' + str(i)] = 'Si' if 'Piscina' in sheet_in[features_col + str(i)].value else 'No'
        sheet_out['W' + str(i)] = 'Si' if sheet_in[cantina_col + str(i)].value in ('Sì','Si') else 'No'
        sheet_out['X' + str(i)] = 'Si' if 'Taverna' in sheet_in[features_col + str(i)].value else 'No'
        if portineria_col is not None:
            sheet_out['Y' + str(i)] = 'Si' if 'Portiere intera giornata' == sheet_in[portineria_col + str(i)].value else 'No'
            sheet_out['Z' + str(i)] = 'Si' if 'Portiere mezza giornata' == sheet_in[portineria_col + str(i)].value else 'No'
        sheet_out['AA' + str(i)] = 'Si' if 'Reception' in sheet_in[features_col + str(i)].value else 'No'
        sheet_out['AB'+ str(i)] = sheet_in[disp_col + str(i)].value
        sheet_out['AC'+ str(i)] = sheet_in[spese_col + str(i)].value
        sheet_out['AD'+ str(i)] = sheet_in[contr_col + str(i)].value
        sheet_out['AE'+ str(i)] = str(sheet_in[proprieta_col + str(i)].value).replace(',','') if sheet_in[proprieta_col + str(i)].value is not None else 'null'
        if warm_type_col is not None:
            sheet_out['AF'+ str(i)] = sheet_in[warm_type_col + str(i)].value
        
        if warm_alim_col is not None:
            sheet_out['AG'+ str(i)] = sheet_in[warm_alim_col + str(i)].value.split(',')[1] if len(sheet_in[warm_alim_col + str(i)].value.split(',')) == 2 else sheet_in[warm_alim_col + str(i)].value if 'alimentato' in sheet_in[warm_alim_col + str(i)].value else 'null'
            sheet_out['AH'+ str(i)] = sheet_in[warm_alim_col + str(i)].value.split(',')[0] if len(sheet_in[warm_alim_col + str(i)].value.split(',')) == 2 else sheet_in[warm_alim_col + str(i)].value if 'alimentato' not in sheet_in[warm_alim_col + str(i)].value else 'null'
        sheet_out['AI'+ str(i)] = sheet_in[class_ene_col + str(i)].value
        sheet_out['AJ'+ str(i)] = sheet_in[ene_eff_col + str(i)].value
        sheet_out['AK' + str(i)] = sheet_in[summer_col + str(i)].value
        sheet_out['AL' + str(i)] = sheet_in[winter_col + str(i)].value
        sheet_out['AM' + str(i)] = strip_words(sheet_in[index_ene_col + str(i)].value)
        infissi = [x for x in sheet_in[features_col + str(i)].value.split(',') if 'Infissi' in x]
        sheet_out['AN' + str(i)] = 'null' if infissi == [] else infissi[0].split('/')[0].replace('Infissi esterni in ','')
        sheet_out['AO' + str(i)] = 'null' if infissi == [] else infissi[0].split('/')[1]
        sheet_out['AP' + str(i)] = 'Si' if 'Caminetto' in sheet_in[features_col + str(i)].value else 'No'
        esposizione = [x for x in sheet_in[features_col + str(i)].value.split(',') if 'Esposizione' in x]
        sheet_out['AQ' + str(i)] = 'null' if esposizione == [] else esposizione[0].replace('Esposizione ','').strip()
        sheet_out['AR' + str(i)] = 'Si' if 'Fibra ottica' in sheet_in[features_col + str(i)].value else 'No'
        
    workbook_out.save(output_path)
    remove_duplicates(output_path)

def remove_duplicates(path):
    wb = load_workbook(path)
    ws = wb.active
    unique_rows= set()

    for row in ws.iter_rows(2, values_only=True):
        unique_rows.add(tuple(row))

    ws.delete_rows(2, ws.max_row)
    for unique_row in unique_rows:
        ws.append(unique_row)
    
    wb.save(path)

if __name__ == '__main__':
    #generate_definitive_table('regioni/Cagliari.xlsx','regioni/Cagliari-Def.xlsx')
    split_into_folders('Roma')